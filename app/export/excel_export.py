# -*- coding: utf-8 -*-
"""DB駆動の帳票生成。scripts/build_dispatch.py のスタイリング・レイアウトを
そのまま踏襲する(過去に一度「見た目が紙帳票から離れすぎた」と差し戻された
経緯があるため、罫線・配色・フォント・列幅は変更しない)。styles.py に
移植済みのスタイル定義を使う。"""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import styles as st
from .. import models

DISPATCH_HEADERS = ["日付", "元請", "現場名", "台数", "自社車番", "運転手", "傭車", "備考"]
INVOICE_HEADERS = ["日付", "名称", "台数", "数量", "単価", "金額", "車番", "備考"]


def build_invoice_workbook(conn, invoice_id):
    """請求書1件分のワークブックを生成する。見た目は build_dispatch.py の
    請求書シートと同一(罫線・配色・列幅・フォント)。金額はinvoicesに固定保存
    済みの値を転記しつつ、明細行はシート内数式(=数量×単価)でも検証できるようにする。"""
    invoice = models.get_invoice(conn, invoice_id)
    if invoice is None:
        raise ValueError(f"invoice {invoice_id} が見つかりません")
    lines = models.list_invoice_lines(conn, invoice_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"
    ws.sheet_view.showGridLines = False
    st.title_cell(ws, "請求書")

    labels = [
        ("A3", "請求先"), ("A4", "請求日"),
        ("D3", "税込合計金額"), ("D4", "消費税10%"),
    ]
    for ref, text in labels:
        ws[ref] = text
        ws[ref].font = Font(bold=True)
        ws[ref].alignment = st.LEFT

    ws["B3"] = invoice["client_name_snapshot"]
    ws["B3"].border = st.BORDER
    ws["B4"] = invoice["invoice_date"]
    ws["B4"].number_format = "yyyy/mm/dd"
    ws["B4"].border = st.BORDER

    ws["E3"] = invoice["total_amount"]
    ws["E3"].number_format = "#,##0円"
    ws["E3"].font = Font(bold=True, size=14, color="C0392B")
    ws["E3"].alignment = st.LEFT

    ws["E4"] = invoice["tax_amount"]
    ws["E4"].number_format = "#,##0円"
    ws["E4"].font = Font(bold=True, size=12)
    ws["E4"].alignment = st.LEFT

    # インボイス制度(適格請求書等保存方式)対応: 発行元の登録番号を明記する
    ws["A5"] = "発行元"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = invoice["issuer_name_snapshot"]
    ws["D5"] = "登録番号"
    ws["D5"].font = Font(bold=True)
    ws["E5"] = invoice["issuer_registration_number_snapshot"] or "(未設定)"

    st.note_cell(ws, "A6", "自動生成された請求書です。内容の追記・修正が必要な場合はこのシートを直接編集してください。")
    st.note_cell(ws, "A7", "金額(F列)は数量×単価で自動計算されます。")

    detail_header_row = 8
    detail_start = 9
    detail_end = detail_start + max(len(lines) + 5, 30) - 1

    for i, h in enumerate(INVOICE_HEADERS, start=1):
        ws.cell(row=detail_header_row, column=i).value = h
    st.style_header_row(ws, detail_header_row, 1, len(INVOICE_HEADERS))
    for i, w in enumerate([12, 24, 8, 8, 10, 12, 14, 26], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for offset, r in enumerate(range(detail_start, detail_end + 1)):
        ws.row_dimensions[r].height = 20
        line = lines[offset] if offset < len(lines) else None

        for col in range(1, 9):
            if col == 6:
                continue
            cell = ws.cell(row=r, column=col)
            cell.border = st.BORDER
            st.input_cell(ws, cell.coordinate)
            cell.alignment = st.CENTER if col in (1, 3, 4, 5) else st.LEFT
        ws.cell(row=r, column=1).number_format = "m/d(aaa)"
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"

        if line is not None:
            ws.cell(row=r, column=1).value = line["entry_date"]
            ws.cell(row=r, column=2).value = line["display_name"]
            ws.cell(row=r, column=3).value = line["count"]
            ws.cell(row=r, column=4).value = line["quantity"]
            ws.cell(row=r, column=5).value = line["unit_price"]
            ws.cell(row=r, column=7).value = line["vehicle_no"]
            ws.cell(row=r, column=8).value = line["memo"]

        f_cell = ws.cell(row=r, column=6)
        f_cell.value = f"=IF(AND($D{r}<>\"\",$E{r}<>\"\"),$D{r}*$E{r},\"\")"
        f_cell.number_format = "#,##0"
        f_cell.border = st.BORDER
        st.linked_cell(ws, f_cell.coordinate)
        f_cell.alignment = st.CENTER

    ws.freeze_panes = f"A{detail_start}"
    ws.print_area = f"A1:H{detail_end}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = st.BUTTON_COLORS["請求書"]

    return wb


def build_dispatch_day_workbook(conn, date_str):
    """指定日の配車表を印刷用に1シートで生成する(会長・現場が紙で見る用、
    scripts/build_dispatch.py の配車入力/印刷シートと同じ見た目)。"""
    entries = [
        e for e in models.list_dispatch_entries(conn, entry_date=date_str)
        if e["client_id"] is not None or e["count"] is not None
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "配車表"
    ws.sheet_view.showGridLines = False
    st.title_cell(ws, f"配車表(印刷用) {date_str}")

    header_row = 4
    data_start = 5
    data_end = data_start + max(len(entries), 1) - 1

    for i, h in enumerate(DISPATCH_HEADERS, start=1):
        ws.cell(row=header_row, column=i).value = h
    st.style_header_row(ws, header_row, 1, len(DISPATCH_HEADERS))
    for i, w in enumerate([12, 20, 24, 8, 14, 10, 20, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for offset, r in enumerate(range(data_start, data_end + 1)):
        ws.row_dimensions[r].height = 20
        entry = entries[offset] if offset < len(entries) else None
        for col in range(1, len(DISPATCH_HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = st.BORDER
            cell.alignment = st.CENTER if col in (1, 4, 6) else st.LEFT
        if entry is not None:
            ws.cell(row=r, column=1).value = entry["entry_date"]
            ws.cell(row=r, column=1).number_format = "m/d(aaa)"
            ws.cell(row=r, column=2).value = entry["client_name_snapshot"]
            ws.cell(row=r, column=3).value = entry["site_name_snapshot"]
            ws.cell(row=r, column=4).value = entry["count"]
            ws.cell(row=r, column=5).value = entry["vehicle_no"]
            ws.cell(row=r, column=6).value = entry["driver_name"]
            ws.cell(row=r, column=7).value = (
                entry["subcontractor_name"] if entry["is_subcontractor"] else "自社"
            )
            ws.cell(row=r, column=8).value = entry["memo"]

    ws.freeze_panes = f"A{data_start}"
    ws.print_area = f"A1:H{data_end}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = st.BUTTON_COLORS["印刷"]

    return wb
