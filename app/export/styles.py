# -*- coding: utf-8 -*-
"""scripts/build_dispatch.py のスタイル定義・セル装飾ヘルパーをそのまま移植したもの。
請求書・配車表の『見た目』は過去に一度差し戻された経緯があるため、
このモジュールでは見た目に関わる値(色・フォント・罫線)は一切変更しない。"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

YELLOW = PatternFill("solid", fgColor="FFFF00")          # 入力セル
GRAY = PatternFill("solid", fgColor="F2F2F2")             # 自動転記(数式)セル
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
NOTE_FONT = Font(size=9, italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

BUTTON_COLORS = {
    "配車入力": "2E86C1",
    "請求用配車": "27AE60",
    "自社売上票": "AF601A",
    "単価表": "8E44AD",
    "請求書": "C0392B",
    "印刷": "16A085",
}


def style_header_row(ws, row, col_start, col_end, height=26):
    ws.row_dimensions[row].height = height
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def back_to_main_button(ws, cell_ref="H1"):
    ws[cell_ref] = "◀ メインへ戻る"
    ws[cell_ref].hyperlink = "#'メイン'!A1"
    ws[cell_ref].font = Font(bold=True, size=12, color="FFFFFF")
    ws[cell_ref].fill = PatternFill("solid", fgColor="555555")
    ws[cell_ref].alignment = CENTER
    ws[cell_ref].border = BORDER


def title_cell(ws, text, cell_ref="A1"):
    ws[cell_ref] = text
    ws[cell_ref].font = TITLE_FONT


def note_cell(ws, cell_ref, text):
    ws[cell_ref] = text
    ws[cell_ref].font = NOTE_FONT


def input_cell(ws, ref, border=True):
    ws[ref].fill = YELLOW
    if border:
        ws[ref].border = BORDER


def linked_cell(ws, ref, border=True):
    ws[ref].fill = GRAY
    ws[ref].font = Font(color="595959")
    if border:
        ws[ref].border = BORDER
