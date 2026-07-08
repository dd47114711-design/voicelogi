# -*- coding: utf-8 -*-
"""
トラック配車システム_Excel版.xlsx を生成するスクリプト。
仕様書(チャット指示)に基づき、写真・既存ファイルなしでゼロから構築。
ハイパーリンクによるシート間ナビゲーション方式(VBAマクロなし)。
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

OUT_PATH = "/tmp/claude-0/-home-user-voicelogi/58cd87a9-9cb3-5ae1-b9a5-8edec09f8a7a/scratchpad/トラック配車システム_Excel版.xlsx"

# 既存の「単価表(20260417)」から抽出した実データ(会社名, 昼, 夜, その他)。
# 価格欄に日付・電話メモ等の文字列が混在している行はそのまま文字列として保持し、
# 誤った数値変換をしないようにしている。
with open("/tmp/claude-0/-home-user-voicelogi/58cd87a9-9cb3-5ae1-b9a5-8edec09f8a7a/scratchpad/price_data.json", encoding="utf-8") as f:
    PRICE_DATA = json.load(f)
N_COMPANIES = len(PRICE_DATA)
MASTER_ROWS = max(120, N_COMPANIES + 20)  # 実データ86社+今後の追加余地

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---------- 共通スタイル ----------
YELLOW = PatternFill("solid", fgColor="FFFF00")          # 入力セル
GRAY = PatternFill("solid", fgColor="F2F2F2")             # 自動転記(数式)セル
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
NOTE_FONT = Font(size=9, italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

BUTTON_COLORS = {
    "配車入力": "2E86C1",
    "請求用配車": "27AE60",
    "自社売上票": "AF601A",
    "単価表": "8E44AD",
    "請求書": "C0392B",
    "印刷": "16A085",
}


def style_header_row(ws, row, col_start, col_end, height=26):
    ws.row_dimensions[row].height = height
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def back_to_main_button(ws, cell_ref="H1"):
    ws[cell_ref] = "◀ メインへ戻る"
    ws[cell_ref].hyperlink = "#'メイン'!A1"
    ws[cell_ref].font = Font(bold=True, size=12, color="FFFFFF")
    ws[cell_ref].fill = PatternFill("solid", fgColor="555555")
    ws[cell_ref].alignment = CENTER
    ws[cell_ref].border = BORDER


def title_cell(ws, text, cell_ref="A1"):
    ws[cell_ref] = text
    ws[cell_ref].font = TITLE_FONT


def note_cell(ws, cell_ref, text):
    ws[cell_ref] = text
    ws[cell_ref].font = NOTE_FONT


def input_cell(ws, ref, border=True):
    ws[ref].fill = YELLOW
    if border:
        ws[ref].border = BORDER


def linked_cell(ws, ref, border=True):
    ws[ref].fill = GRAY
    ws[ref].font = Font(color="595959")
    if border:
        ws[ref].border = BORDER


# =====================================================================
# 1. 隠しマスタ・裏集計シート
# =====================================================================

# --- 元請マスタ(既存の単価表から実在86社を投入) ---
ws = wb.create_sheet("元請マスタ")
ws["A1"] = "元請名"; ws["B1"] = "備考"
style_header_row(ws, 1, 1, 2, 20)
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 30
for r in range(2, MASTER_ROWS + 2):
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER
for i, (name, hiru, yoru, sonota) in enumerate(PRICE_DATA):
    ws.cell(row=2 + i, column=1).value = name

# --- 地名マスタ(現場名) ---
ws = wb.create_sheet("地名マスタ")
ws["A1"] = "現場名"; ws["B1"] = "関連元請"; ws["C1"] = "備考"
style_header_row(ws, 1, 1, 3, 20)
for col, w in zip("ABC", (26, 24, 30)):
    ws.column_dimensions[col].width = w
for r in range(2, 202):
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER

# --- 車両マスタ(自社車両) ---
ws = wb.create_sheet("車両マスタ")
ws["A1"] = "自社車番"; ws["B1"] = "車種"; ws["C1"] = "既定運転手"; ws["D1"] = "車検日"
style_header_row(ws, 1, 1, 4, 20)
for col, w in zip("ABCD", (16, 14, 16, 14)):
    ws.column_dimensions[col].width = w
for r in range(2, 102):
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BORDER
ws.cell(row=r, column=4).number_format = "yyyy/mm/dd"
for r in range(2, 102):
    ws.cell(row=r, column=4).number_format = "yyyy/mm/dd"

# --- 運転手マスタ(配車入力の写真にあった実在18名を投入) ---
DRIVER_ROSTER = ["坂本", "松興", "興栄", "山田", "竹本", "中村", "金子", "フジタ", "永井",
                  "久富", "水谷", "渡辺", "小田", "青木", "中本", "木内", "山住", "西本"]
ws = wb.create_sheet("運転手マスタ")
ws["A1"] = "運転手名"; ws["B1"] = "電話番号"; ws["C1"] = "備考"
style_header_row(ws, 1, 1, 3, 20)
for col, w in zip("ABC", (16, 18, 30)):
    ws.column_dimensions[col].width = w
for r in range(2, 102):
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER
for i, name in enumerate(DRIVER_ROSTER):
    ws.cell(row=2 + i, column=1).value = name

# --- 傭車マスタ ---
ws = wb.create_sheet("傭車マスタ")
ws["A1"] = "傭車会社名"; ws["B1"] = "連絡先"; ws["C1"] = "備考"
style_header_row(ws, 1, 1, 3, 20)
for col, w in zip("ABC", (22, 18, 30)):
    ws.column_dimensions[col].width = w
for r in range(2, 102):
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER

# --- 単価マスタ(単価表を自動展開した参照用一覧) ---
# 実際の単価表を確認したところ、4tは列ではなく「会社名 4t車」という別行で
# 管理されていたため、区分は 昼/夜/その他 の3種類とし、4t車は単価表に
# 別行(例:「〇〇建設 4t車」)として登録する運用にしている。
ws = wb.create_sheet("単価マスタ")
ws["A1"] = "会社名"; ws["B1"] = "区分"; ws["C1"] = "単価"; ws["D1"] = "適用開始日"
style_header_row(ws, 1, 1, 4, 20)
note_cell(ws, "F1", "※単価表シートの内容を自動展開した一覧です。直接編集しないでください。")
CATS = ["昼", "夜", "その他"]
CAT_COL = {"昼": "B", "夜": "C", "その他": "D"}
r_out = 2
for i in range(MASTER_ROWS):
    src_row = i + 5  # 単価表のデータ開始行(5)に合わせる
    for cat in CATS:
        ws.cell(row=r_out, column=1).value = f"=IF('単価表'!$A${src_row}=\"\",\"\",'単価表'!$A${src_row})"
        ws.cell(row=r_out, column=2).value = cat
        ws.cell(row=r_out, column=3).value = f"=IF('単価表'!${CAT_COL[cat]}${src_row}=\"\",\"\",'単価表'!${CAT_COL[cat]}${src_row})"
        ws.cell(row=r_out, column=4).value = f"=IF('単価表'!$E${src_row}=\"\",\"\",'単価表'!$E${src_row})"
        r_out += 1
for col, w in zip("ABCD", (22, 10, 12, 14)):
    ws.column_dimensions[col].width = w

# --- 旧配車表(将来の過去データ保管用プレースホルダ) ---
ws = wb.create_sheet("旧配車表")
ws["A1"] = "旧配車表(過去データ保管用・非表示シート)"
ws["A1"].font = Font(bold=True, size=12)
note_cell(ws, "A2", "旧システムからの過去データを移行する場合はこのシートに貼り付けてください。")

# --- 裏集計(選択リスト・検証用) ---
ws = wb.create_sheet("裏集計")
ws["A1"] = "区分リスト"
for i, v in enumerate(CATS, start=2):
    ws.cell(row=i, column=1).value = v
ws["B1"] = "傭車区分リスト"
for i, v in enumerate(["自社", "傭車"], start=2):
    ws.cell(row=i, column=2).value = v
ws["C1"] = "チェックリスト"
for i, v in enumerate(["済", "未"], start=2):
    ws.cell(row=i, column=3).value = v
note_cell(ws, "E1", "各シートのドロップダウンの選択肢を管理しています。行の追加・削除はしないでください。")

print("マスタ・裏集計シート作成完了")

# =====================================================================
# 2. 定義名(ドロップダウン用リスト)
# =====================================================================
def add_defined_name(name, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=ref)

add_defined_name("元請リスト", f"元請マスタ!$A$2:$A${MASTER_ROWS + 1}")
add_defined_name("現場リスト", "地名マスタ!$A$2:$A$201")
add_defined_name("車番リスト", "車両マスタ!$A$2:$A$101")
add_defined_name("運転手リスト", "運転手マスタ!$A$2:$A$101")
add_defined_name("傭車リスト", "傭車マスタ!$A$2:$A$101")
add_defined_name("区分リスト", f"裏集計!$A$2:$A${1 + len(CATS)}")
add_defined_name("傭車区分リスト", "裏集計!$B$2:$B$3")
add_defined_name("チェックリスト", "裏集計!$C$2:$C$3")

print("定義名作成完了")

# =====================================================================
# 3. 単価表(表示シート)
# =====================================================================
ws = wb.create_sheet("単価表")
ws.sheet_view.showGridLines = False
title_cell(ws, "単価表")
back_to_main_button(ws, "H1")
note_cell(ws, "A2", "会社ごとに 昼/夜/その他 の単価を管理します。4t車は別行として「会社名　4t車」のように登録してください(既存の単価表の運用にあわせています)。適用開始日を分ければ単価改定にも対応できます。")
note_cell(ws, "A3", f"既存の単価表(2026/4/17時点)から実在{N_COMPANIES}社を自動転記しました。日付・電話番号等のメモが混在していた価格欄はそのまま文字列で転記しているため、数値に直す場合はセルを直接編集してください。")

PRICE_DATA_START = 5
PRICE_DATA_END = PRICE_DATA_START + MASTER_ROWS - 1

headers = ["会社名", "昼", "夜", "その他", "適用開始日", "備考"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i).value = h
style_header_row(ws, 4, 1, len(headers))
widths = [26, 14, 14, 16, 14, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(PRICE_DATA_START, PRICE_DATA_END + 1):
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        input_cell(ws, cell.coordinate)
        if c in (2, 3, 4):
            cell.number_format = "#,##0"
        elif c == 5:
            cell.number_format = "yyyy/mm/dd"

# 実データ投入(会社名, 昼, 夜, その他)
for i, (name, hiru, yoru, sonota) in enumerate(PRICE_DATA):
    r = PRICE_DATA_START + i
    ws.cell(row=r, column=1).value = name
    ws.cell(row=r, column=2).value = hiru
    ws.cell(row=r, column=3).value = yoru
    ws.cell(row=r, column=4).value = sonota
    # 文字列(日付・電話メモ混在)が入った価格セルは数値書式を外す
    for col, val in ((2, hiru), (3, yoru), (4, sonota)):
        if isinstance(val, str):
            ws.cell(row=r, column=col).number_format = "General"

ws.freeze_panes = "A5"
ws.print_area = f"A1:F{PRICE_DATA_END}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.tabColor = BUTTON_COLORS["単価表"]

print("単価表シート作成完了")

# =====================================================================
# 4. 配車入力シート(前日に配車を組む・紙の配車表に近い横長レイアウト)
# =====================================================================
ws = wb.create_sheet("配車入力")
ws.sheet_view.showGridLines = False
title_cell(ws, "配車入力(前日配車表)")
back_to_main_button(ws, "I1")
note_cell(ws, "A2", "黄色いセルに入力してください。ここに入力すると『請求用配車』シートに自動で反映されます。")

DISPATCH_HEADERS = ["日付", "元請", "現場名", "台数", "自社車番", "傭車", "傭車名", "備考"]
HEADER_ROW = 4
DATA_START = 5
DATA_END = 204  # 200行分

for i, h in enumerate(DISPATCH_HEADERS, start=1):
    ws.cell(row=HEADER_ROW, column=i).value = h
style_header_row(ws, HEADER_ROW, 1, len(DISPATCH_HEADERS))
widths = [12, 20, 24, 8, 14, 10, 20, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(DATA_START, DATA_END + 1):
    ws.row_dimensions[r].height = 20
    for c in range(1, len(DISPATCH_HEADERS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        input_cell(ws, cell.coordinate)
        cell.alignment = CENTER if c in (1, 4, 6) else LEFT
    ws.cell(row=r, column=1).number_format = "m/d(aaa)"
    ws.cell(row=r, column=4).number_format = "0"

# データ検証(ドロップダウン)
def add_dv(ws, formula, cell_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv

add_dv(ws, "=元請リスト", f"B{DATA_START}:B{DATA_END}")
add_dv(ws, "=現場リスト", f"C{DATA_START}:C{DATA_END}")
add_dv(ws, "=車番リスト", f"E{DATA_START}:E{DATA_END}")
add_dv(ws, "=傭車区分リスト", f"F{DATA_START}:F{DATA_END}")
add_dv(ws, "=傭車リスト", f"G{DATA_START}:G{DATA_END}")

# 傭車=傭車のときだけ傭車名列を目立たせる(条件付き書式)
ws.conditional_formatting.add(
    f"G{DATA_START}:G{DATA_END}",
    FormulaRule(formula=[f"$F{DATA_START}=\"自社\""], fill=PatternFill("solid", fgColor="D9D9D9")),
)

# --- 出勤チェックパネル(K〜L列。紙の配車表にあった運転手一覧+○×チェック欄) ---
note_cell(ws, "K2", "その日出勤している運転手を○/×でチェックできます(運転手マスタと連動)。")
ws["K3"] = "出勤チェック(対象日)"
ws["K3"].font = Font(bold=True, size=11)
ws["L3"] = ""
input_cell(ws, "L3"); ws["L3"].border = BORDER
ws["L3"].number_format = "m/d(aaa)"

ROSTER_HEADER_ROW = 4
ROSTER_START = 5
ROSTER_END = ROSTER_START + len(DRIVER_ROSTER) - 1
ws.cell(row=ROSTER_HEADER_ROW, column=11).value = "運転手名"
ws.cell(row=ROSTER_HEADER_ROW, column=12).value = "出勤"
style_header_row(ws, ROSTER_HEADER_ROW, 11, 12, height=20)
ws.column_dimensions["K"].width = 14
ws.column_dimensions["L"].width = 10

for i in range(len(DRIVER_ROSTER)):
    r = ROSTER_START + i
    name_cell = ws.cell(row=r, column=11)
    name_cell.value = f"=IF(運転手マスタ!A{2 + i}=\"\",\"\",運転手マスタ!A{2 + i})"
    name_cell.border = BORDER
    linked_cell(ws, name_cell.coordinate)
    name_cell.alignment = LEFT

    chk_cell = ws.cell(row=r, column=12)
    chk_cell.border = BORDER
    input_cell(ws, chk_cell.coordinate)
    chk_cell.alignment = CENTER

dv_roster = DataValidation(type="list", formula1='"○,×"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_roster)
dv_roster.add(f"L{ROSTER_START}:L{ROSTER_END}")

ws.freeze_panes = f"A{DATA_START}"
ws.print_area = f"A1:H{DATA_END}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.tabColor = BUTTON_COLORS["配車入力"]

print("配車入力シート作成完了")

# =====================================================================
# 5. 請求用配車シート(実績整理・請求書のもとになる表)
# =====================================================================
ws = wb.create_sheet("請求用配車")
ws.sheet_view.showGridLines = False
title_cell(ws, "請求用配車(実績整理)")
back_to_main_button(ws, "N1")
note_cell(ws, "A2", "日付〜傭車名・備考は『配車入力』から自動転記されます(灰色セル)。運転手・区分・数量H・チェックは黄色セルに入力してください。単価は単価表から自動計算されます。4t車の場合は『配車入力』の元請欄で「会社名 4t車」を選んでください。")
note_cell(ws, "A3", "右側の『請求書貼付用』(N〜T列)は、フィルタで元請・チェック=済を絞り込んだ後、請求書シートへコピー貼り付けするための整形済みデータです。")

BILL_HEADERS = ["日付", "元請", "現場名", "台数", "自社車番", "運転手", "傭車名",
                "区分(昼/夜/その他)", "数量H", "単価", "金額", "チェック", "備考"]
HEADER_ROW = 4
DATA_START = 5
DATA_END = 204

for i, h in enumerate(BILL_HEADERS, start=1):
    ws.cell(row=HEADER_ROW, column=i).value = h
style_header_row(ws, HEADER_ROW, 1, len(BILL_HEADERS))
widths = [12, 20, 24, 8, 14, 14, 20, 18, 10, 10, 12, 10, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 請求書貼付用ヘッダー(N〜T列)
PASTE_HEADERS = ["日付", "名称(現場名)", "台数", "数量", "単価", "車番", "備考"]
for i, h in enumerate(PASTE_HEADERS, start=14):  # N=14
    c = ws.cell(row=HEADER_ROW, column=i)
    c.value = h
style_header_row(ws, HEADER_ROW, 14, 14 + len(PASTE_HEADERS) - 1)
note_cell(ws, "N3", "▼請求書貼付用(フィルタ後にコピー)")
paste_widths = [12, 22, 8, 8, 10, 14, 26]
for i, w in enumerate(paste_widths, start=14):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(DATA_START, DATA_END + 1):
    dr = r  # 配車入力の対応行(1:1)
    ws.row_dimensions[r].height = 20

    # --- 配車入力から自動転記(灰色・数式) ---
    links = {
        1: f"=IF('配車入力'!A{dr}=\"\",\"\",'配車入力'!A{dr})",   # 日付
        2: f"=IF('配車入力'!B{dr}=\"\",\"\",'配車入力'!B{dr})",   # 元請
        3: f"=IF('配車入力'!C{dr}=\"\",\"\",'配車入力'!C{dr})",   # 現場名
        4: f"=IF('配車入力'!D{dr}=\"\",\"\",'配車入力'!D{dr})",   # 台数
        5: f"=IF('配車入力'!E{dr}=\"\",\"\",'配車入力'!E{dr})",   # 自社車番
        7: f"=IF('配車入力'!G{dr}=\"\",\"\",'配車入力'!G{dr})",   # 傭車名
        13: f"=IF('配車入力'!H{dr}=\"\",\"\",'配車入力'!H{dr})",  # 備考
    }
    for col, formula in links.items():
        cell = ws.cell(row=r, column=col)
        cell.value = formula
        cell.border = BORDER
        linked_cell(ws, cell.coordinate)
        if col == 1:
            cell.number_format = "m/d(aaa)"
        cell.alignment = CENTER if col in (1, 4) else LEFT

    # --- 手入力(黄色) ---
    for col in (6, 8, 9, 12):  # 運転手, 区分, 数量H, チェック
        cell = ws.cell(row=r, column=col)
        cell.border = BORDER
        input_cell(ws, cell.coordinate)
        cell.alignment = CENTER if col in (8, 9, 12) else LEFT
    ws.cell(row=r, column=9).number_format = "0.0"

    # --- 単価(単価表から自動計算) ---
    price_cell = ws.cell(row=r, column=10)
    price_cell.value = (
        f"=IFERROR(INDEX('単価表'!$B${PRICE_DATA_START}:$D${PRICE_DATA_END},"
        f"MATCH($B{r},'単価表'!$A${PRICE_DATA_START}:$A${PRICE_DATA_END},0),"
        f"MATCH($H{r},{{\"昼\",\"夜\",\"その他\"}},0)),\"\")"
    )
    price_cell.number_format = "#,##0"
    price_cell.border = BORDER
    linked_cell(ws, price_cell.coordinate)
    price_cell.alignment = CENTER

    # --- 金額(数量H×単価) ---
    amount_cell = ws.cell(row=r, column=11)
    amount_cell.value = f"=IF(AND($I{r}<>\"\",$J{r}<>\"\"),$I{r}*$J{r},\"\")"
    amount_cell.number_format = "#,##0"
    amount_cell.border = BORDER
    linked_cell(ws, amount_cell.coordinate)
    amount_cell.alignment = CENTER

    # --- 請求書貼付用(N〜T列) ---
    paste_map = {
        14: f"=A{r}",          # 日付
        15: f"=C{r}",          # 名称(現場名)
        16: f"=D{r}",          # 台数
        17: f"=I{r}",          # 数量
        18: f"=J{r}",          # 単価
        19: f"=IF(E{r}<>\"\",E{r},G{r})",  # 車番(自社車番、無ければ傭車名)
        20: f"=M{r}",          # 備考
    }
    for col, formula in paste_map.items():
        cell = ws.cell(row=r, column=col)
        cell.value = formula
        cell.border = BORDER
        cell.alignment = CENTER if col in (14, 16, 17, 18) else LEFT
        if col == 14:
            cell.number_format = "m/d(aaa)"
        if col in (17, 18):
            cell.number_format = "#,##0"

# チェック=済の行を緑でハイライト
ws.conditional_formatting.add(
    f"A{DATA_START}:M{DATA_END}",
    FormulaRule(formula=[f"$L{DATA_START}=\"済\""], fill=PatternFill("solid", fgColor="D5F5E3")),
)

add_dv(ws, "=運転手リスト", f"F{DATA_START}:F{DATA_END}")
add_dv(ws, "=区分リスト", f"H{DATA_START}:H{DATA_END}")
add_dv(ws, "=チェックリスト", f"L{DATA_START}:L{DATA_END}")

ws.freeze_panes = f"A{DATA_START}"
ws.auto_filter.ref = f"A{HEADER_ROW}:M{DATA_END}"
ws.print_area = f"A1:M{DATA_END}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.tabColor = BUTTON_COLORS["請求用配車"]

print("請求用配車シート作成完了")

# =====================================================================
# 6. 請求書シート
# =====================================================================
ws = wb.create_sheet("請求書")
ws.sheet_view.showGridLines = False
title_cell(ws, "請求書")
back_to_main_button(ws, "G1")

# --- ヘッダー情報 ---
# 明細は 8行目=ヘッダー, 9〜68行目=データ(60行), 金額列=F
DETAIL_HEADER_ROW = 8
DETAIL_START = 9
DETAIL_END = 68
SUBTOTAL_RANGE = f"F{DETAIL_START}:F{DETAIL_END}"

labels = [
    ("A3", "請求先"), ("A4", "請求日"),
    ("D3", "税込合計金額"), ("D4", "消費税10%"),
]
for ref, text in labels:
    ws[ref] = text
    ws[ref].font = Font(bold=True)
    ws[ref].alignment = LEFT

ws["B3"] = ""
input_cell(ws, "B3"); ws["B3"].border = BORDER
add_dv(ws, "=元請リスト", "B3:B3")
ws["B4"] = ""
input_cell(ws, "B4"); ws["B4"].border = BORDER
ws["B4"].number_format = "yyyy/mm/dd"

ws["E3"] = f"=ROUND(SUM({SUBTOTAL_RANGE})*1.1,0)"   # 税込合計金額
ws["E3"].number_format = "#,##0円"
ws["E3"].font = Font(bold=True, size=14, color="C0392B")
ws["E3"].alignment = LEFT

ws["E4"] = f"=ROUND(SUM({SUBTOTAL_RANGE})*0.1,0)"    # 消費税10%
ws["E4"].number_format = "#,##0円"
ws["E4"].font = Font(bold=True, size=12)
ws["E4"].alignment = LEFT

note_cell(ws, "A6", "同じ名称が続く場合は自動的に「〃」と表示されます。金額=数量×単価で自動計算されます(数量は基本「台数×8H」を入力)。")
note_cell(ws, "A7", "A〜H列はすべて自動表示です。直接入力せず、J〜P列(貼付欄)に入力するか、『請求用配車』シートでフィルタした行(N〜T列)をコピーしてJ列へ値貼り付けしてください。")

# --- 明細ヘッダー(表示専用・すべて数式) ---
INVOICE_HEADERS = ["日付", "名称", "台数", "数量", "単価", "金額", "車番", "備考"]
for i, h in enumerate(INVOICE_HEADERS, start=1):
    ws.cell(row=DETAIL_HEADER_ROW, column=i).value = h
style_header_row(ws, DETAIL_HEADER_ROW, 1, len(INVOICE_HEADERS))
widths = [12, 24, 8, 8, 10, 12, 14, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- J〜P列: 貼付用の生データ(黄色・請求用配車のN〜T列と同じ並び) ---
PASTE_ZONE_HEADERS = ["日付", "名称", "台数", "数量", "単価", "車番", "備考"]
for i, h in enumerate(PASTE_ZONE_HEADERS, start=10):  # J=10
    ws.cell(row=DETAIL_HEADER_ROW, column=i).value = h
style_header_row(ws, DETAIL_HEADER_ROW, 10, 10 + len(PASTE_ZONE_HEADERS) - 1)
note_cell(ws, "J7", "▼貼付欄(請求用配車のN〜T列をそのまま貼り付け)")
paste_zone_widths = [12, 22, 8, 8, 10, 14, 26]
for i, w in enumerate(paste_zone_widths, start=10):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(DETAIL_START, DETAIL_END + 1):
    ws.row_dimensions[r].height = 20

    # J〜P: 生データ入力(黄色)
    for col in range(10, 17):  # J..P
        cell = ws.cell(row=r, column=col)
        cell.border = BORDER
        input_cell(ws, cell.coordinate)
        cell.alignment = CENTER if col in (10, 12, 13, 14) else LEFT
    ws.cell(row=r, column=10).number_format = "m/d(aaa)"   # J 日付
    ws.cell(row=r, column=13).number_format = "#,##0"       # M 数量
    ws.cell(row=r, column=14).number_format = "#,##0"       # N 単価

    # A 日付 = J
    a_cell = ws.cell(row=r, column=1)
    a_cell.value = f"=IF($J{r}=\"\",\"\",$J{r})"
    a_cell.number_format = "m/d(aaa)"

    # B 名称 = K列参照。前行と同じなら「〃」
    b_cell = ws.cell(row=r, column=2)
    if r == DETAIL_START:
        b_cell.value = f"=IF($K{r}=\"\",\"\",$K{r})"
    else:
        b_cell.value = f"=IF($K{r}=\"\",\"\",IF($K{r}=$K{r-1},\"〃\",$K{r}))"

    # C 台数 = L, D 数量 = M, E 単価 = N
    c_cell = ws.cell(row=r, column=3); c_cell.value = f"=IF($L{r}=\"\",\"\",$L{r})"
    d_cell = ws.cell(row=r, column=4); d_cell.value = f"=IF($M{r}=\"\",\"\",$M{r})"
    e_cell = ws.cell(row=r, column=5); e_cell.value = f"=IF($N{r}=\"\",\"\",$N{r})"
    d_cell.number_format = "#,##0"
    e_cell.number_format = "#,##0"

    # F 金額 = 数量 × 単価 (自動計算)。数量は基本的に「台数×8H」を入力する運用。
    f_cell = ws.cell(row=r, column=6)
    f_cell.value = f"=IF(AND($D{r}<>\"\",$E{r}<>\"\"),$D{r}*$E{r},\"\")"
    f_cell.number_format = "#,##0"

    # G 車番 = O, H 備考 = P
    g_cell = ws.cell(row=r, column=7); g_cell.value = f"=IF($O{r}=\"\",\"\",$O{r})"
    h_cell = ws.cell(row=r, column=8); h_cell.value = f"=IF($P{r}=\"\",\"\",$P{r})"

    for col, cell in ((1, a_cell), (2, b_cell), (3, c_cell), (4, d_cell),
                       (5, e_cell), (6, f_cell), (7, g_cell), (8, h_cell)):
        cell.border = BORDER
        linked_cell(ws, cell.coordinate)
        cell.alignment = CENTER if col in (1, 3, 4, 5, 6) else LEFT

ws.freeze_panes = f"A{DETAIL_START}"
ws.print_area = f"A1:H{DETAIL_END}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.tabColor = BUTTON_COLORS["請求書"]

print("請求書シート作成完了")

# =====================================================================
# 7. 自社売上票シート(自社車両ごとの日別売上管理)
# =====================================================================
ws = wb.create_sheet("自社売上票")
ws.sheet_view.showGridLines = False
title_cell(ws, "自社売上票")

ws["A3"] = "対象月"
ws["A3"].font = Font(bold=True)
ws["B3"] = ""
input_cell(ws, "B3"); ws["B3"].border = BORDER
ws["B3"].number_format = "yyyy年m月"
note_cell(ws, "D3", "日ごとに売上・距離(km)・燃費を入力してください。休みの日は売上欄に「休」と入力すると月合計・日数の計算から自動で除外されます。")

SALES_ROW1 = 5   # 日付番号(3列結合)
SALES_ROW2 = 6   # 売上/距離/燃費 の見出し
SALES_DATA_START = 7
SALES_DATA_END = 26  # 自社車両20台分

DAY_BLOCK_START = 4          # D列から日ブロック開始
SUB_LABELS = ["売上", "距離(km)", "燃費"]
N_DAYS = 31

# 左側の固定列(運転手名・車検日・車番)は2段ヘッダーを縦結合
for col, label, w in ((1, "運転手名", 14), (2, "車検日", 12), (3, "車番", 12)):
    ws.merge_cells(start_row=SALES_ROW1, start_column=col, end_row=SALES_ROW2, end_column=col)
    ws.cell(row=SALES_ROW1, column=col).value = label
    ws.column_dimensions[get_column_letter(col)].width = w

# 日ブロック(31日 × 3列)
for d in range(1, N_DAYS + 1):
    block_start = DAY_BLOCK_START + (d - 1) * 3
    ws.merge_cells(start_row=SALES_ROW1, start_column=block_start,
                    end_row=SALES_ROW1, end_column=block_start + 2)
    ws.cell(row=SALES_ROW1, column=block_start).value = d
    for s, label in enumerate(SUB_LABELS):
        c = ws.cell(row=SALES_ROW2, column=block_start + s)
        c.value = label
        ws.column_dimensions[get_column_letter(block_start + s)].width = 9 if s == 0 else 7

DAY_BLOCK_END = DAY_BLOCK_START + N_DAYS * 3 - 1  # 最終日の燃費列

# 右側の集計列(休/月合計/距離合計/日数/1日平均)も2段ヘッダーを縦結合
SUMMARY_LABELS = ["休", "月合計", "距離合計", "日数", "1日平均"]
SUMMARY_WIDTHS = [7, 12, 12, 7, 10]
summary_cols = list(range(DAY_BLOCK_END + 1, DAY_BLOCK_END + 1 + len(SUMMARY_LABELS)))
for col, label, w in zip(summary_cols, SUMMARY_LABELS, SUMMARY_WIDTHS):
    ws.merge_cells(start_row=SALES_ROW1, start_column=col, end_row=SALES_ROW2, end_column=col)
    ws.cell(row=SALES_ROW1, column=col).value = label
    ws.column_dimensions[get_column_letter(col)].width = w
LAST_COL = summary_cols[-1]

back_to_main_button(ws, f"{get_column_letter(LAST_COL + 2)}1")

style_header_row(ws, SALES_ROW1, 1, LAST_COL, height=20)
style_header_row(ws, SALES_ROW2, 1, LAST_COL, height=18)

sales_cols = [DAY_BLOCK_START + (d - 1) * 3 for d in range(1, N_DAYS + 1)]
distance_cols = [c + 1 for c in sales_cols]

for r in range(SALES_DATA_START, SALES_DATA_END + 1):
    ws.row_dimensions[r].height = 20
    # 運転手名・車検日・車番(入力)
    for col in (1, 2, 3):
        cell = ws.cell(row=r, column=col)
        cell.border = BORDER
        input_cell(ws, cell.coordinate)
    ws.cell(row=r, column=2).number_format = "yyyy/mm/dd"

    for d in range(1, N_DAYS + 1):
        block_start = DAY_BLOCK_START + (d - 1) * 3
        sales_cell = ws.cell(row=r, column=block_start)
        dist_cell = ws.cell(row=r, column=block_start + 1)
        fuel_cell = ws.cell(row=r, column=block_start + 2)
        for cell in (sales_cell, dist_cell, fuel_cell):
            cell.border = BORDER
            input_cell(ws, cell.coordinate)
            cell.alignment = CENTER
        sales_cell.number_format = "#,##0"
        dist_cell.number_format = "0"
        fuel_cell.number_format = "0.00"

    sales_refs = [f"{get_column_letter(c)}{r}" for c in sales_cols]
    dist_refs = [f"{get_column_letter(c)}{r}" for c in distance_cols]

    rest_cell = ws.cell(row=r, column=summary_cols[0])
    rest_cell.value = "=" + "+".join(f'COUNTIF({ref},"休")' for ref in sales_refs)

    total_cell = ws.cell(row=r, column=summary_cols[1])
    total_cell.value = "=SUM(" + ",".join(sales_refs) + ")"
    total_cell.number_format = "#,##0"

    dist_total_cell = ws.cell(row=r, column=summary_cols[2])
    dist_total_cell.value = "=SUM(" + ",".join(dist_refs) + ")"
    dist_total_cell.number_format = "#,##0"

    count_cell = ws.cell(row=r, column=summary_cols[3])
    count_cell.value = "=COUNT(" + ",".join(sales_refs) + ")"

    avg_cell = ws.cell(row=r, column=summary_cols[4])
    avg_cell.value = f"=IFERROR({total_cell.coordinate}/{count_cell.coordinate},0)"
    avg_cell.number_format = "#,##0"

    for cell in (rest_cell, total_cell, dist_total_cell, count_cell, avg_cell):
        cell.border = BORDER
        linked_cell(ws, cell.coordinate)
        cell.alignment = CENTER

add_dv(ws, "=運転手リスト", f"A{SALES_DATA_START}:A{SALES_DATA_END}")
add_dv(ws, "=車番リスト", f"C{SALES_DATA_START}:C{SALES_DATA_END}")

ws.freeze_panes = f"D{SALES_DATA_START}"
ws.print_area = f"A1:{get_column_letter(LAST_COL)}{SALES_DATA_END}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.tabColor = BUTTON_COLORS["自社売上票"]

print("自社売上票シート作成完了")

# =====================================================================
# 8. 印刷シート(配車入力を印刷用に整えたコピー)
# =====================================================================
ws = wb.create_sheet("印刷")
ws.sheet_view.showGridLines = False
title_cell(ws, "配車表(印刷用)")
back_to_main_button(ws, "I1")
note_cell(ws, "A2", "『配車入力』の内容がそのまま反映されます。オートフィルタで日付を絞り込んでから印刷してください(Ctrl+P)。")

PRINT_HEADERS = DISPATCH_HEADERS
P_HEADER_ROW = 4
P_DATA_START = 5
P_DATA_END = 204

for i, h in enumerate(PRINT_HEADERS, start=1):
    ws.cell(row=P_HEADER_ROW, column=i).value = h
style_header_row(ws, P_HEADER_ROW, 1, len(PRINT_HEADERS))
widths = [12, 20, 24, 8, 14, 10, 20, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(P_DATA_START, P_DATA_END + 1):
    dr = r
    ws.row_dimensions[r].height = 20
    for col in range(1, len(PRINT_HEADERS) + 1):
        src = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=IF('配車入力'!{src}{dr}=\"\",\"\",'配車入力'!{src}{dr})"
        cell.border = BORDER
        cell.alignment = CENTER if col in (1, 4, 6) else LEFT
        if col == 1:
            cell.number_format = "m/d(aaa)"

ws.freeze_panes = f"A{P_DATA_START}"
ws.auto_filter.ref = f"A{P_HEADER_ROW}:H{P_DATA_END}"
ws.print_area = f"A1:H{P_DATA_END}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = f"{P_HEADER_ROW}:{P_HEADER_ROW}"
ws.sheet_properties.tabColor = BUTTON_COLORS["印刷"]

print("印刷シート作成完了")

# =====================================================================
# 9. メインシート(ナビゲーション・最初に開く画面)
# =====================================================================
ws = wb.create_sheet("メイン", 0)
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:M3")
ws["B2"] = "トラック配車システム"
ws["B2"].font = Font(bold=True, size=26, color="1F4E78")
ws["B2"].alignment = CENTER

ws.merge_cells("B4:M4")
ws["B4"] = "使いたい画面の大きなボタンをクリックしてください"
ws["B4"].font = Font(size=13, color="595959")
ws["B4"].alignment = CENTER
ws.row_dimensions[4].height = 22

BUTTON_LAYOUT = [
    ("配車入力", "明日の配車を組む(毎日の入力はここから)"),
    ("請求用配車", "実績を整理する(請求書のもと)"),
    ("自社売上票", "自社車両の日別売上を管理する"),
    ("単価表", "会社ごとの単価(昼/夜/その他)を管理する"),
    ("請求書", "請求書を作成する"),
    ("印刷", "配車表を印刷する"),
]

BTN_COL_SPANS = [("B", "F"), ("H", "L")]   # 2列
BTN_ROW_HEIGHT = 6                          # 行数(高さ)
BTN_ROW_GAP = 1
START_ROW = 6

from openpyxl.utils import column_index_from_string

for idx, (name, desc) in enumerate(BUTTON_LAYOUT):
    col_pair = BTN_COL_SPANS[idx % 2]
    row_block = idx // 2
    top_row = START_ROW + row_block * (BTN_ROW_HEIGHT + BTN_ROW_GAP + 2)
    bottom_row = top_row + BTN_ROW_HEIGHT - 1
    ref = f"{col_pair[0]}{top_row}:{col_pair[1]}{bottom_row}"
    ws.merge_cells(ref)
    cell = ws[f"{col_pair[0]}{top_row}"]
    cell.value = name
    cell.hyperlink = f"#'{name}'!A1"
    cell.font = Font(bold=True, size=20, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BUTTON_COLORS[name])
    cell.alignment = CENTER
    cell.border = Border(left=Side(style="medium", color="FFFFFF"),
                          right=Side(style="medium", color="FFFFFF"),
                          top=Side(style="medium", color="FFFFFF"),
                          bottom=Side(style="medium", color="FFFFFF"))
    for rr in range(top_row, bottom_row + 1):
        ws.row_dimensions[rr].height = 24

    desc_row = bottom_row + 1
    desc_ref = f"{col_pair[0]}{desc_row}:{col_pair[1]}{desc_row}"
    ws.merge_cells(desc_ref)
    dcell = ws[f"{col_pair[0]}{desc_row}"]
    dcell.value = desc
    dcell.font = Font(size=10, color="808080")
    dcell.alignment = CENTER
    ws.row_dimensions[desc_row].height = 16

for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
    ws.column_dimensions[col].width = 11
ws.column_dimensions["G"].width = 4

last_bottom = START_ROW + (len(BUTTON_LAYOUT) // 2 - 1) * (BTN_ROW_HEIGHT + BTN_ROW_GAP + 2) + BTN_ROW_HEIGHT + 2
note_cell(ws, f"B{last_bottom + 1}", "※マスタ(車両・運転手・元請・地名・傭車・単価)や旧配車表・裏集計は非表示シートに保管されています。通常操作では触る必要はありません。")

ws.sheet_view.zoomScale = 100
ws.sheet_properties.tabColor = "1F4E78"
wb.active = wb.sheetnames.index("メイン")

print("メインシート作成完了")

# =====================================================================
# 10. シート表示順・表示/非表示の最終設定
# =====================================================================
VISIBLE_ORDER = ["メイン", "配車入力", "請求用配車", "自社売上票", "単価表", "請求書", "印刷"]
HIDDEN_SHEETS = ["車両マスタ", "運転手マスタ", "元請マスタ", "地名マスタ",
                 "傭車マスタ", "単価マスタ", "旧配車表", "裏集計"]

ordered = [wb[name] for name in VISIBLE_ORDER] + [wb[name] for name in HIDDEN_SHEETS]
wb._sheets = ordered

for name in HIDDEN_SHEETS:
    wb[name].sheet_state = "hidden"
for name in VISIBLE_ORDER:
    wb[name].sheet_state = "visible"

wb.active = wb.sheetnames.index("メイン")
wb.calculation.fullCalcOnLoad = True

wb.save(OUT_PATH)
print("saved (final)")
print("sheets:", [(s.title, s.sheet_state) for s in wb._sheets])

